import json
from openpyxl import Workbook

def generate_excel_report(json_file, excel_file):
    # Открываем файл JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)['data']

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Заголовки столбцов
    headers = ["Оператор", "Количество чатов", "Пунктуація", "Простота розмови", "Рейтинг", "Оценка диалога", "Понимание клиента"]

    # Запись заголовков
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Запись данных
    for row_num, row_data in enumerate(data, 2):
        chat_count = len(row_data["chats"])  # Получаем количество чатов
        ws.cell(row=row_num, column=1, value=row_data["operator"])
        ws.cell(row=row_num, column=2, value=chat_count)  # Записываем количество чатов
        # Рассчитываем среднее значение пунктуации, простоты речи, рейтинга, оценки диалога и понимания клиента
        avg_punctuation = row_data["correct_punctuation"] / chat_count if chat_count > 0 else 0
        avg_simplicity = row_data["simplicity_of_language"] / chat_count if chat_count > 0 else 0
        avg_rating = row_data["operator_rating"] / chat_count if chat_count > 0 else 0
        avg_dialogue_assessment = row_data["dialogue_assessment"] / chat_count if chat_count > 0 else 0
        avg_customer_insight = row_data["customer_insight"] / chat_count if chat_count > 0 else 0
        ws.cell(row=row_num, column=3, value=avg_punctuation)
        ws.cell(row=row_num, column=4, value=avg_simplicity)
        ws.cell(row=row_num, column=5, value=avg_rating)
        ws.cell(row=row_num, column=6, value=avg_dialogue_assessment)
        ws.cell(row=row_num, column=7, value=avg_customer_insight)

    # Сохраняем книгу Excel
    wb.save(excel_file)


